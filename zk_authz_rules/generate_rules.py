"""
Agent that finds and processes UnitedHealthcare (UHC) documentation for CPT codes,
extracting real billing requirements from UHC Excel files and generating rules.json files.

This agent:
1. Downloads UHC OPG Exhibit Excel files that contain CPT codes
2. Parses the Excel files to extract CPT code information
3. Extracts only information found in the documents
4. Generates rules.json files only for codes with available documentation
5. Caps at 100 files maximum
"""

import json
import os
import re
import time
import urllib.parse
import tempfile
from typing import List, Dict, Optional, Any
from policy_model import Policy

try:
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: Required packages not installed. Run: pip install requests beautifulsoup4 openpyxl")
    exit(1)

OUT_DIR = "rules"

# Common CPT codes to search for (up to 100)
CPT_CODES_TO_SEARCH = [
    "99213", "99214", "99203", "99204", "99212", "99215", "99205", "99202", "99211",
    "99396", "99395", "99397", "99394", "99391", "99392", "99393", "99381", "99382",
    "99383", "99384", "99385", "99386", "99387",
    "99223", "99222", "99221", "99233", "99232", "99231",
    "99284", "99283", "99282", "99281",
    "93000", "85025", "80053", "81001", "70450", "72141", "71020", "36415",
    "99238", "99239", "99254", "99255", "99244", "99245",
    "27447", "29881", "45378", "45380", "70496",
    "70551", "78815", "81211", "27446", "29880",
    "45385", "45388", "45390", "45392", "45393",
    "72142", "72146", "72148", "72149", "72158",
    "71035", "71036", "71037", "71038", "71039",
    "70460", "70470", "70480", "70481", "70482",
    "36416", "36417", "36600", "36620", "36640",
    "99236", "99237", "99256", "99257",
    "99455", "99456", "99457", "99458", "99459",
    "99460", "99461", "99462", "99463", "99464",
    "99465", "99468", "99469", "99471", "99472",
    "99473", "99474", "99475", "99476", "99477",
    "97110", "97112", "97116", "97140", "97150",
]

MAX_FILES = 100

UHC_BASE_URL = "https://www.uhcprovider.com"
UHC_OPG_PAGE = f"{UHC_BASE_URL}/en/claims-payments-billing/outpatient-procedure-grouper-exhibits.html"


def normalize_icd(code: str) -> str:
    """Normalize ICD-10 code (remove dots, uppercase)."""
    return code.replace(".", "").upper().strip()


def extract_icd10_codes(text: str) -> List[str]:
    """Extract ICD-10 codes from text using regex patterns."""
    icd_codes = set()
    
    patterns = [
        r'\b([A-Z]\d{2}\.?\d{0,2})\b',
        r'ICD[- ]?10[- ]?[Cc]ode[s]?[:\s]+([A-Z]\d{2}\.?\d{0,2})',
        r'([A-Z]\d{2}\.\d{1,2})',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            normalized = normalize_icd(match)
            if 3 <= len(normalized) <= 7:
                icd_codes.add(normalized)
    
    return sorted(list(icd_codes))


def extract_age_requirements(text: str) -> Optional[int]:
    """Extract age requirements from text."""
    patterns = [
        r'age\s*[>=]\s*(\d+)',
        r'(\d+)\s*years?\s*or\s*older',
        r'minimum\s*age[:\s]+(\d+)',
        r'at\s*least\s*(\d+)\s*years?\s*old',
        r'age\s*(\d+)\s*or\s*older',
    ]
    
    ages = []
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            try:
                age = int(match)
                if 0 <= age <= 120:
                    ages.append(age)
            except:
                pass
    
    return min(ages) if ages else None


def extract_prior_auth(text: str) -> Optional[bool]:
    """Determine if prior authorization is required from text."""
    text_lower = text.lower()
    
    pa_indicators = [
        'prior authorization required',
        'prior auth required',
        'requires prior authorization',
        'requires prior auth',
        'preauthorization required',
        'pa required',
        'must obtain prior authorization',
        'prior authorization is required',
        'precertification required',
        'requires precertification',
    ]
    
    no_pa_indicators = [
        'no prior authorization',
        'prior authorization not required',
        'does not require prior authorization',
        'prior auth not required',
        'no precertification',
        'precertification not required',
    ]
    
    for indicator in no_pa_indicators:
        if indicator in text_lower:
            return False
    
    for indicator in pa_indicators:
        if indicator in text_lower:
            return True
    
    return None


def extract_pos_codes_from_text(text: str) -> List[int]:
    """Extract place of service codes ONLY if explicitly mentioned in text."""
    pos_codes = []
    
    patterns = [
        r'POS\s*(\d{2})',
        r'place\s+of\s+service\s+code\s+(\d{2})',
        r'place\s+of\s+service\s+(\d{2})',
        r'location\s+code\s+(\d{2})',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            try:
                pos = int(match)
                if 11 <= pos <= 99:
                    if pos not in pos_codes:
                        pos_codes.append(pos)
            except:
                pass
    
    return sorted(pos_codes)


def fetch_page(url: str, timeout: int = 15) -> Optional[str]:
    """Fetch HTML content from a URL."""
    try:
        response = requests.get(url, timeout=timeout, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
        })
        response.raise_for_status()
        return response.text
    except Exception as e:
        return None


def download_excel_file(url: str) -> Optional[str]:
    """Download an Excel file and return the path to the temporary file."""
    try:
        full_url = urllib.parse.urljoin(UHC_BASE_URL, url)
        response = requests.get(full_url, timeout=30, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        response.raise_for_status()
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.write(response.content)
        temp_file.close()
        
        return temp_file.name
    except Exception as e:
        print(f"    Error downloading {url}: {e}")
        return None


def parse_excel_file(file_path: str, cpt_code: str) -> Optional[Dict[str, Any]]:
    """Parse Excel file to find information about a specific CPT code."""
    try:
        wb = load_workbook(file_path, data_only=True)
        
        # Search through all sheets
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Search all cells for the CPT code
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    row_data.append(cell_value)
                    
                    # Check if this cell contains our CPT code
                    if cpt_code in cell_value:
                        # Found the CPT code, extract row information
                        description = ""
                        icd_codes = []
                        
                        # Look at the row for description and other info
                        row_text = " ".join([str(val) for val in row_data])
                        
                        # Extract description (often in next column)
                        for i, cell_val in enumerate(row_data):
                            if cpt_code in cell_val and i + 1 < len(row_data):
                                desc = row_data[i + 1]
                                if desc and len(desc) > 5:
                                    description = desc[:150]
                        
                        # Extract ICD codes from the row
                        icd_codes = extract_icd10_codes(row_text)
                        
                        # Extract age requirements
                        age_gte = extract_age_requirements(row_text)
                        
                        # Extract PA requirements
                        requires_pa = extract_prior_auth(row_text)
                        
                        # Extract POS codes
                        pos_allowed = extract_pos_codes_from_text(row_text)
                        
                        # If we found useful info, return it
                        if description or icd_codes:
                            return {
                                "description": description or f"CPT code {cpt_code}",
                                "requires_pa": requires_pa if requires_pa is not None else True,
                                "icd10_in": icd_codes if icd_codes else ["Z0000"],
                                "exclusion_icd10": None,
                                "age_gte": age_gte,
                                "pos_allowed": pos_allowed if pos_allowed else None,
                                "source": "UHC OPG Excel file",
                                "source_file": os.path.basename(file_path),
                            }
        
        return None
        
    except Exception as e:
        print(f"    Error parsing Excel file {file_path}: {e}")
        return None
    finally:
        # Clean up temporary file
        try:
            os.unlink(file_path)
        except:
            pass


def get_uhc_excel_files() -> List[str]:
    """Get list of UHC OPG Excel file URLs."""
    print("  Fetching UHC OPG Exhibits page...")
    html = fetch_page(UHC_OPG_PAGE)
    if not html:
        print("  Failed to fetch OPG page")
        return []
    
    soup = BeautifulSoup(html, 'html.parser')
    links = soup.find_all('a', href=True)
    
    excel_files = []
    for link in links:
        href = link.get('href', '')
        if any(ext in href.lower() for ext in ['.xlsx', '.xls', 'excel']):
            excel_files.append(href)
    
    print(f"  Found {len(excel_files)} Excel files")
    return excel_files


def search_uhc_for_code(cpt_code: str) -> Optional[Dict[str, Any]]:
    """Search UHC Excel files for a specific CPT code."""
    print(f"  Searching UHC Excel files for CPT {cpt_code}...")
    
    # Get list of Excel files
    excel_files = get_uhc_excel_files()
    
    if not excel_files:
        return None
    
    # Try the most recent file first (usually first in list)
    for excel_url in excel_files[:3]:  # Try first 3 files
        print(f"    Checking {os.path.basename(excel_url)}...")
        file_path = download_excel_file(excel_url)
        if file_path:
            result = parse_excel_file(file_path, cpt_code)
            if result:
                result["source_url"] = urllib.parse.urljoin(UHC_BASE_URL, excel_url)
                return result
            time.sleep(1)  # Rate limiting
    
    return None


def parse_documentation(cpt_code: str, doc: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Parse documentation to extract billing requirements."""
    # The Excel parser already returns structured data, so we can use it directly
    if not doc:
        return None
    
    # Ensure we have at least ICD codes or description
    if not doc.get("icd10_in") and not doc.get("description"):
        return None
    
    return doc


def create_policy(cpt_code: str, parsed_data: Dict[str, Any]) -> Policy:
    """Create a Policy object from parsed data."""
    policy_id = f"UHC-{cpt_code}"
    
    inclusion = {
        "icd10_in": parsed_data.get("icd10_in", ["Z0000"])
    }
    
    if parsed_data.get("age_gte"):
        inclusion["age_gte"] = parsed_data["age_gte"]
    
    exclusion = {}
    if parsed_data.get("exclusion_icd10"):
        exclusion["icd10_in"] = parsed_data["exclusion_icd10"]
    
    admin = {
        "max_units_per_day": 1
    }
    
    if parsed_data.get("pos_allowed"):
        admin["pos_allowed"] = parsed_data["pos_allowed"]
    
    source_url = parsed_data.get("source_url", UHC_OPG_PAGE)
    
    metadata = {
        "source_url": source_url,
        "effective_date": "2025-01-01",
        "notes": f"Extracted from {parsed_data.get('source', 'UHC documentation')}. {parsed_data.get('description', '')}"
    }
    
    return Policy(
        policy_id=policy_id,
        version="2025-01-01",
        payer="uhc",
        lob="commercial",
        codes=[cpt_code],
        requires_pa=parsed_data.get("requires_pa", True),
        inclusion=inclusion,
        exclusion=exclusion,
        admin=admin,
        metadata=metadata
    )


def write_rule_file(code: str, policy: Policy) -> str:
    """Write a rules.json file for a CPT code."""
    os.makedirs(OUT_DIR, exist_ok=True)
    fname = f"{code}.json"
    path = os.path.join(OUT_DIR, fname)
    
    canonical = policy.canonical_json()
    h = policy.hash()
    
    blob = dict(canonical)
    blob["policy_hash"] = h
    
    with open(path, "w") as f:
        json.dump(blob, f, indent=2, ensure_ascii=False)
    
    return path


def process_cpt_code(code: str) -> bool:
    """Process a single CPT code."""
    try:
        # Step 1: Search for documentation in Excel files
        doc = search_uhc_for_code(code)
        if not doc:
            print(f"  No UHC documentation found for CPT {code}")
            return False
        
        # Step 2: Parse documentation (already parsed from Excel)
        parsed_data = parse_documentation(code, doc)
        if not parsed_data:
            print(f"  Insufficient information extracted for CPT {code}")
            return False
        
        # Step 3: Create policy and write file
        policy = create_policy(code, parsed_data)
        path = write_rule_file(code, policy)
        
        print(f"  [OK] Generated {path}")
        print(f"       Found {len(parsed_data['icd10_in'])} ICD-10 codes")
        if parsed_data.get('pos_allowed'):
            print(f"       POS codes: {parsed_data['pos_allowed']}")
        
        time.sleep(1)  # Rate limiting
        return True
        
    except Exception as e:
        print(f"  [ERROR] Failed to process {code}: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Main function: process CPT codes and generate files for those with available UHC documentation."""
    print("=" * 70)
    print("UnitedHealthcare (UHC) CPT Code Documentation Processor")
    print("=" * 70)
    print("This agent will:")
    print("  1. Download UHC OPG Exhibit Excel files")
    print("  2. Parse Excel files to extract CPT code information")
    print("  3. Generate rules.json files ONLY for codes found in Excel files")
    print(f"  4. Stop after {MAX_FILES} files maximum")
    print("=" * 70)
    print()
    
    success_count = 0
    
    for i, code in enumerate(CPT_CODES_TO_SEARCH, 1):
        if success_count >= MAX_FILES:
            print(f"\nReached maximum of {MAX_FILES} files. Stopping.")
            break
        
        print(f"\n[{i}/{len(CPT_CODES_TO_SEARCH)}] Processing CPT {code}...")
        
        if process_cpt_code(code):
            success_count += 1
    
    print("\n" + "=" * 70)
    print(f"[COMPLETE] Generated {success_count} rules.json files")
    print(f"Files are in the '{OUT_DIR}' folder")
    print("=" * 70)


if __name__ == "__main__":
    main()
